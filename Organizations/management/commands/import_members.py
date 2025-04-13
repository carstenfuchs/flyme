from datetime import datetime
from dateutil.parser import parse
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from Organizations.models import Membership


def normalize(datapoints):
    new_datapoints = []
    not_handled = {}

    for dp in datapoints:
        new_dp = {}

        for key, value in dp.items():
            if value is None:
                continue

            if key in ("Anrede", "Fax", "Konto-Nr", "BLZ", "KI", "J-Beitr. 1", "J-Beitr. 2", "Jahresbeitrag", "Lastschrift"):
                continue

            if key in ("ref_date", "Vorname", "Nachname", "Straße", "PLZ", "Ort"):
                new_dp[key] = value
                continue

            if key == "Name":
                new_dp["Nachname"] = value
                continue

            if key == "Name, Vorname":
                n, v = value.split(",")
                new_dp["Vorname"] = v.strip()
                new_dp["Nachname"] = n.strip()
                continue

            if key in ("Telefon", "Festnetz", "Handy"):
                if "Telefon" not in new_dp:
                    new_dp["Telefon"] = []
                v = value.replace("/", " ").replace("-", " ")
                # Remove all but the first " ".
                v = v.replace(" ", "*", 1).replace(" ", "").replace("*", " ")
                if not v.startswith("0"):
                    v = "0" + v
                new_dp["Telefon"].append(v)
                continue

            if key in ("E-Mail", "Mail"):
                new_dp["E-Mail"] = value
                continue

            if key in ("Geburt", "Geburtstag"):
                if isinstance(value, datetime):
                    new_dp["Geburtstag"] = value
                else:
                    new_dp["Geburtstag"] = parse(value)
                continue

            if key in ("Eintritt", "Aufnahme"):
                new_dp["Aufnahme"] = value # parse(value)
                continue

            if key in ("Status",):
                v = value.lower()
                if v == "a" or "aktiv" in v:
                    new_dp["Status"] = Membership.STATUS_ACTIVE
                elif v in ("p", "f") or "passiv" in v:
                    new_dp["Status"] = Membership.STATUS_SUPPORTING
                elif v == "e" or "ehren" in v:
                    new_dp["Status"] = Membership.STATUS_HONORARY
                else:
                    new_dp["Status"] = Membership.STATUS_OTHER
                continue

            # The (key, value) pair was not handled above.
            if key not in not_handled:
                print(f"Not handled: „{key}“ (e.g. „{value}“)")
                not_handled[key] = value
            new_dp[key] = value

        # Skip empty rows.
        if not "Nachname" in new_dp:
            continue

        # Normalize even further.
        v = new_dp.pop("Vorname", "").strip()
        n = new_dp.pop("Nachname", "").strip()
        new_dp["short_name"] = v or n or "Mitglied"
        new_dp["full_name"] = f"{v} {n}"
        new_dp["sort_name"] = f"{n}, {v}"

        s = new_dp.pop("Straße", "").strip()
        p = str(new_dp.pop("PLZ", "")).strip()
        o = new_dp.pop("Ort", "").strip()
        new_dp["postal_address"] = f"{s}\n{p} {o}".strip()

        print(new_dp)
        new_datapoints.append(new_dp)

    return new_datapoints


def integrate(org, members, datapoints):
    for dp in datapoints:
        key = f"{dp['full_name']} ({dp['Geburtstag']})"

        if not key in members:
            u = User(
                full_name=dp['full_name'],
                short_name=dp['short_name'],
              # sort_name=dp['sort_name'],
                email=dp['E-Mail'],     # !!! What if none?
                password='Test',
            )

            ms = Membership(
                user=u,
                orga=org,
                status=xy,
                begin=xy,
                end=xy,
                remark="",
            )

            members[key] = (u, [ms])

        u, list_ms = members[key]
        # ...


    # Not each member in `members` may have been mentioned in `datapoints`.
    # For these members, we have to add a new datapoint with "resigned" status
    # – unless the last datapoint has a "resigned" status already.
    pass


class Command(BaseCommand):
    help = "Loads membership data from xlsx files."

    def add_arguments(self, parser):
        # See https://stackoverflow.com/questions/15753701/how-can-i-pass-a-list-as-a-command-line-argument-with-argparse
        parser.add_argument("org", action="store", type=int, help="The ID of the organization that the memberships are loaded into.")
        parser.add_argument("infile", nargs='+', help="One or more files to read membership data from.")
        parser.add_argument("--update-database", action="store_true", help="The database is not modified without this flag being set.")

    def handle(self, *args, **options):
        # This dict maps "full_name (day_of_birth)" strings to tuples
        # (User instance, list of chronologically ordered Membership instances).
        all_members = {}

        for fn in options['infile']:
            self.stdout.write(f"{fn}")
            wb = load_workbook(fn, data_only=True)
            ws = wb.active

            ref_date = ws["B1"].value.date()
            print(f"worksheet: max_row {ws.max_row}, max_column {ws.max_column}, ref_date at B1: {ref_date}")

            num_empty_rows = 0
            last_row_nr = 3

            for row in ws.iter_rows(min_col=1, max_col=1, min_row=4):
                cell = row[0]
                if not cell.value:
                    num_empty_rows += 1
                    if num_empty_rows > 5:
                        break
                    continue
                last_row_nr = cell.row

            # Each datapoint represents the member data as given in the input file.
            # The datapoints at indices 0, 1 and 2 (the header rows) will be unused.
            datapoints = [{"ref_date": ref_date} for i in range(last_row_nr)]
            print("last_row_nr", last_row_nr)

            for col in ws.columns:
                key = col[2].value
                if not key:
                    break
                print(key)

                for cell in col[3:]:
                    if cell.row > last_row_nr:
                        break
                    # print(cell.row, cell.value)
                    # Also record empty strings and `None` values:
                    datapoints[cell.row - 1][key] = cell.value

            # All rows in this file were loaded into `datapoints`.
            # Now integrate them into the large members dict.
            datapoints = normalize(datapoints)
            integrate(all_members, datapoints)

        if options["update_database"]:
            # ....
            self.stdout.write(self.style.SUCCESS("Database updated."))
